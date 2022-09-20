from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "jaewonsheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 의  셀의 정보를 출력
print(ws["A1"].value) # A1 셀의 '값' 을 출력

print(ws["A10"].value) # 값이 없을때는 none 을 출력



wb.save("sample.xlsx")
