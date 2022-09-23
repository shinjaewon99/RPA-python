from openpyxl import Workbook

wb = Workbook()
# wb.active
ws = wb.create_sheet() #새로운 sheet 기본 이름으로 생성

ws.title = "Mysheet"  # 시트 이름 변경
ws.sheet_properties.tabColor ="ff66ff" # RGB 형태로 입력


ws1 = wb.create_sheet("Yoursheet") # 주어진 이름으로 시트 생성
ws2 = wb.create_sheet("Newsheet" , 2) # 2번째 index 에 시트 생성

new_ws = wb["Newsheet"] # Dict 형태로 접근 

print(wb.sheetnames)  # 모든 시트 이름 확인

# 시트 복사

new_ws ["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"


wb.save("sample.xlsx")



