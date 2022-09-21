import imp
from openpyxl import Workbook
from random import *

wb = Workbook()

ws = wb.active


# 1줄씩 데이터 넣기

ws.append(["번호", "영어" , "수학"])

for i in range(1, 11): # 10개 데이터 넣기
    ws.append([i, randint(0,100), randint(0,100)])


col_B = ws["B"] # 영어 컬럼만 가져오기

print(col_B)

for cell in col_B:
    print(cell.value)

col_range = ws["B:C"] # 영어 , 수학  컬럼 함께 가져오기

for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title =ws[1] # 1 번째 row 만 가져오기
for cell in row_title:
    print(cell.value)


row_range = ws[2:6] # 1번째 줄인 title 을 제외하고 2번째에서 6번째 줄까지 가져오기
for rows in row_range:
    for cell in rows:
        print(cell.value , end= " ")
    print()


from openpyxl.utils.cell import coordinate_from_string

row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
for rows in row_range:
    for cell in rows:
        print(cell.coordinate, end="") #A/10 , AZ/250
        xy = coordinate_from_string(cell.coordinate)
        # print(xy, end=" ")
        print(xy[0] , end=" ") # A
        print(xy[1] , end=" ") # 1

    print()


# 전체 row
print(tuple(ws.rows))

# 전체 columns

print(tuple(ws.columns))

wb.save("sample.xlsx")

